# Fortigate Config Parser


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import re
import ipaddress
from openpyxl import Workbook


def parse_address(lines):
    addr_list = []
    current_name = None
    for line in lines:
        line = line.strip()
        m = re.match(r'^edit\s+"([^"]+)"$', line)
        if m:
            current_name = m.group(1)
            continue
        m = re.match(r'^set subnet ([\d\.]+) ([\d\.]+)', line)
        if m and current_name:
            ip, mask = m.group(1), m.group(2)
            try:
                net = ipaddress.IPv4Network(f"{ip}/{mask}", strict=False)
                addr_list.append((current_name, str(net.network_address)))
            except Exception:
                addr_list.append((current_name, ip))
        if line == 'next':
            current_name = None
    return addr_list


def parse_addrgrp(lines, addr_dict):
    groups = []
    current_group = None
    for line in lines:
        line = line.strip()
        m = re.match(r'^edit\s+"([^"]+)"$', line)
        if m:
            current_group = m.group(1)
            continue
        m = re.match(r'^set member (.+)$', line)
        if m and current_group:
            for mbr in re.findall(r'"([^"]+)"', m.group(1)):
                groups.append((current_group, mbr, addr_dict.get(mbr, 'N/A')))
        if line == 'next':
            current_group = None
    return groups


def parse_policy(lines):
    policies = []
    current = None
    for line in lines:
        line = line.strip()
        if line.startswith('edit '):
            if current:
                policies.append(current)
            current = {
                'id': line.split()[1],
                'name': '',
                'srcintf': [],
                'dstintf': [],
                'srcaddr': [],
                'dstaddr': [],
                'service': [],
                'internet_service': False,
                'internet_service_name': [],
                'action': '',
                'status': ''
            }
            continue
        if not current:
            continue
        if line.startswith('set name '):
            current['name'] = re.findall(r'"(.*?)"', line)[0]
        elif line.startswith('set srcintf '):
            current['srcintf'] = re.findall(r'"(.*?)"', line)
        elif line.startswith('set dstintf '):
            current['dstintf'] = re.findall(r'"(.*?)"', line)
        elif line.startswith('set srcaddr '):
            current['srcaddr'] = re.findall(r'"(.*?)"', line)
        elif line.startswith('set dstaddr '):
            current['dstaddr'] = re.findall(r'"(.*?)"', line)
        elif line.startswith('set service '):
            current['service'].extend(re.findall(r'"(.*?)"', line))
        elif line.startswith('set internet-service '):
            current['internet_service'] = ('enable' in line)
        elif line.startswith('set internet-service-name '):
            current['internet_service_name'] = re.findall(r'"(.*?)"', line)
        elif line.startswith('set action '):
            current['action'] = line.split()[2]
        elif line.startswith('set status '):
            current['status'] = line.split()[2]
    if current:
        policies.append(current)
    return policies


def main(input_file, output_file):
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except Exception as e:
        print(f"[Error] 無法讀取檔案: {e}")
        sys.exit(1)

    addr_lines, grp_lines, pol_lines = [], [], []
    mode = None
    for line in lines:
        t = line.strip()
        if t == 'config firewall address': mode = 'addr'; continue
        if t == 'config firewall addrgrp': mode = 'grp'; continue
        if t == 'config firewall policy': mode = 'pol'; continue
        if t == 'end': mode = None; continue
        if mode == 'addr': addr_lines.append(line)
        if mode == 'grp': grp_lines.append(line)
        if mode == 'pol': pol_lines.append(line)

    addr_list = parse_address(addr_lines)
    addr_dict = {n: ip for n, ip in addr_list}
    grp_list = parse_addrgrp(grp_lines, addr_dict)
    pol_list = parse_policy(pol_lines)

    wb = Workbook()
    wb.remove(wb.active)

    # address 分頁
    ws = wb.create_sheet('address')
    ws.append(['名稱', 'IP'])
    for name, ip in addr_list:
        ws.append([name, ip])

    # group 分頁
    ws = wb.create_sheet('group')
    ws.append(['群組', '成員', 'IP'])
    for grp, mbr, ip in grp_list:
        ws.append([grp, mbr, ip])

    # policy 分頁
    ws = wb.create_sheet('policy')
    ws.append(['ID', '名稱', '來源介面', '目的介面', '來源地址', '目的地址', '服務', '動作', '狀態'])
    for p in pol_list:
        # 判斷目的和服務欄位
        if not p['dstaddr'] and p['internet_service']:
            dst = ','.join(p['internet_service_name'])
            srv = ['網際網路服務']
        else:
            dst = ','.join(p.get('dstaddr', []))
            srv = p.get('service', [])
        ws.append([
            p.get('id', ''),
            p.get('name', ''),
            ','.join(p.get('srcintf', [])),
            ','.join(p.get('dstintf', [])),
            ','.join(p.get('srcaddr', [])),
            dst,
            ','.join(srv),
            p.get('action', ''),
            p.get('status', '')
        ])

    try:
        wb.save(output_file)
        print(f"[✓] 完成：{output_file}")
    except Exception as e:
        print(f"[Error] 無法寫入 Excel: {e}")


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("用法: python fortips.py 備份.conf 輸出.xlsx")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])

